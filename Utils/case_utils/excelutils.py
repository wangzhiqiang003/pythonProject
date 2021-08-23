import openpyxl


class ExcelUtils:
    def __init__(self, path):
        self.book = self.create_book(path)

    # 读取每一列的信息
    def readall_by_sheet_name(self, sheet_name):
        sheet = self.book[sheet_name]

        maxrow = sheet.max_row
        maxclomn = sheet.max_column
        list_result = []
        keys = [sheet.cell(row=1, column=x).value for x in range(1, maxclomn + 1)]
        for i in range(2, maxrow + 1):
            values = [sheet.cell(row=i, column=x).value for x in range(1, maxclomn + 1)]
            values = list(map(lambda x: x if x else {}, values))
            list_result.append(dict(zip(keys, values)))
            # print(list_result)

        return list_result

    # 将所有的表信息整合成json
    def assemble_info(self):
        res = {}
        for sheet_name in self.book.sheetnames:
            res[sheet_name] = self.readall_by_sheet_name(sheet_name)
        if res['case']:
            temp = res.pop('case')
            temp = list(filter(lambda x: x['isValide'] == 'Y', temp))
            res['case'] = temp

        return res

    def create_book(self, path):
        return openpyxl.load_workbook(path)


if __name__ == '__main__':
    e = ExcelUtils('../../API_CASE_INFO.xlsx')
    print(e.assemble_info())
